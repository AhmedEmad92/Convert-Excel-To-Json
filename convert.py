from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# import io
import sys
import json



sheet = input ("Please input sheet Name ")
if sheet[-3:] != "xls" and  sheet[-4:] != "xlsx":
    
    try:
        fname = sheet + ".xlsx"
        wb = load_workbook(filename=fname)
        ws = wb.active
    except :
        try:
            fname = sheet+ ".xls"
            wb = load_workbook(filename=fname)
            ws = wb.active
        except: 
            print("Excel not found ")
            sys.exit()

else:
    try:
        wb = load_workbook(filename=sheet)
        ws = wb.active
    except: 
        print("Excel not found ")
        sys.exit()
json_file = input("please input the output file name ")
if json_file[-4:] != "json":
    json_file = json_file + ".json"

my_list = []

last_column = len(list(ws.columns))
last_row = len(list(ws.rows))

for row in range(1, last_row + 1):
    my_dict = {}
    for column in range(1, last_column + 1):
        column_letter = get_column_letter(column)
        if row > 1:
            val = ws[column_letter + str(row)].value
            if val != None :
                my_dict[ws[column_letter + str(1)].value] = val
    my_list.append(my_dict)

data = json.dumps(my_list, sort_keys=True, indent=4,default=str,ensure_ascii=False)

with open(json_file, 'w', encoding='utf-8') as f:
    f.write(data)
f.close()